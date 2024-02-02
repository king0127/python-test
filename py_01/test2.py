
import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0)

data = [
    ('k', 'i', 'h', 'a'),
    ('k1', 'i1', 'h1', 'a1')
]

title_date = ('A', 'B', 'C', 'D')

data.insert(0, title_date)

i=1
for line in data:
    print("输出行数据", line)
    for x in range(0, len(line)):
        print(x, "测试数据", len(line), line[x])
        ws.cell(column= x+1, row=i, value =  "%s" % line[x])
    i += 1

wb.save('./test2.xls')
