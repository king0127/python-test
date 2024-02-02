import openpyxl


# 判断数据是否为函数，剔除为空的数值
def valChange(sth):
    if sth == None:
        sth = 0
    return sth


lw = openpyxl.load_workbook('test4.xlsx')

print('文件基础信息',lw)

table = lw.active
print(table)
max_rows = table.max_row
max_cols =  table.max_column

print(max_cols, max_rows)


total_list=[]
for i in range(max_rows):
    row_list = []
    for line in range(max_cols):
        # print('打印数据', line)
        if i >= 0:
            val = table.cell(i+1, line+1).value
            row_list.append(val)
            print(val, end=" ")
    total_list.append(row_list)
    print('\t')


print(total_list)




