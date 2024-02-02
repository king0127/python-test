import openpyxl

def save_data(data, out_file_name):
    if not out_file_name.endswith('.xls'):
        out_file_name += '.xls'

    wb = openpyxl.Workbook()
    # ws = wb.active 默认第一个sheet

    ws = wb.create_sheet(title='人员信息', index=0)

    i=1
    for line in data:
        for x in range(0, len(line)):
            ws.cell(column=x+1, row=i, value= "%s" % line[x])
        i += 1

    wb.save(out_file_name)



data = [
    ('序号', '名称', '电话', '性别'),
    ('1', 'king', '10086', '男'),
    ('2', 'Tom', '10086', '男'),
    ('3', 'Jonles', '10086', '男'),
]

save_data(data, './test4.xls')








