import openpyxl
import re
import old_mysql_util

pattern = r'"&(\w+)&"'
# 判断数据是否为函数，剔除为空的数值
def valChange(sth):
    if sth == None:
        sth = 0
    return sth

# 去掉函数符号=""
def remove(value):
    if value and isinstance(value, str) and value.startswith('='):
       value = value[2:-1]  # 去掉=“”部分
    return value

def replaceDyc(val, _col_val):
    if val and isinstance(val, str):
        val = re.sub(pattern, _col_val, val)
    return val


lw = openpyxl.load_workbook('D:\wangshuai95012_58016\Desktop\数据迁移验证.xlsx')

print('文件基础信息',lw)

# sheet = lw['2']

# print(lw.worksheets)
lw.sheetnames

table = lw.active
# print(table)
max_rows = table.max_row
max_cols =  table.max_column
# print(max_cols, max_rows)


total_list=[]
for i in range(max_rows):
    row_list = []
    for line in range(max_cols):
        # print('打印数据', line)
        if i >= 0:
            val = table.cell(i+1, line+1).value
            row_list.append(val)
            # print(val, end=" ")
    total_list.append(row_list)
    # print('\t')

# print(total_list)


for data in total_list:
    # print(i)
    project_name = data[0]
    count_sql = replaceDyc(remove(data[1]), project_name)
    num_sql = replaceDyc(remove(data[2]), project_name)
    # print(project_name, count_sql, num_sql)
    # print(num_sql)
    if count_sql != '数量验证':
        res_count = old_mysql_util.querySchedule(count_sql)
        print(res_count)
    if num_sql != '关键字验证':
        res_num = old_mysql_util.querySchedule(num_sql)
        print(res_num)
    # print(project_name, count_sql, num_sql)
    # print(res_count)

# 设定计划ID
schedule_id = 122134

# 查询计划的基础信息




# 文件对比

